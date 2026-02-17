"""Pytest fixtures and helpers for creating sample Excel data."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest


# Column headers required by excel_reader (order can vary; we use a fixed order for tests)
HEADERS = [
    "Teilnehmyliste",
    "Teilnehmyliste E-Mail",
    "Teilnehmyliste Telefonnummer",
    "Teilnehmyliste Nachname",
    "Teilnehmerliste Vorname",
    "Teilnehmyliste Bild",
    "Land",
    "Rufname/Pseudonym",
    "Teilnehmyliste_Couch",
    "E-Mail Adresse",
    "Telefonnummer (mit Ländercode!)",
    "Familiename",
    "Vorname",
]


def build_sample_xlsx(
    tmp_path: Path,
    rows: list[dict[str, object]],
    filename: str = "sample.xlsx",
) -> Path:
    """
    Create a minimal .xlsx with HEADERS in row 1 and data rows.
    Each row dict keys must match HEADERS; missing keys become empty cells.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("No active sheet")
    for col, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)
    for row_idx, row_data in enumerate(rows, start=2):
        for col, header in enumerate(HEADERS, start=1):
            value = row_data.get(header)
            ws.cell(row=row_idx, column=col, value=value)
    out = tmp_path / filename
    wb.save(out)
    return out


@pytest.fixture
def placeholder_path(tmp_path: Path) -> Path:
    """A minimal placeholder image (1x1 PNG) for tests."""
    from PIL import Image
    png = tmp_path / "placeholder.png"
    img = Image.new("RGB", (1, 1), color=(0, 0, 0))
    img.save(png)
    return png
